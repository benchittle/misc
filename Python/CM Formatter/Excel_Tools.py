import openpyxl

class Selection(openpyxl.worksheet.cell_range.CellRange):#openpyxl.worksheet.worksheet.Worksheet):
    def __init__(self, sheet, min_col=None, min_row=None, max_col=None, max_row=None):
        """
        Represents a selection of cells.

        Arguments:
        sheet -- openpyxl Worksheet object to contain the selection
        min_col -- int
        min_row -- int
        max_col -- int
        max_row -- int
        """

        openpyxl.worksheet.cell_range.CellRange.__init__(self, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)

        if type(sheet) != openpyxl.worksheet.worksheet.Worksheet:
            raise TypeError("'sheet' must be a valid Worksheet object")

        self.sheet = sheet

        self.height = self.max_row - self.min_row
        self.width =  self.max_col - self.min_col


    def getContents(self, as_list=True, by_row=False):
        """
        Return a list or generator containing the contents of the cells within the selection.

        Arguments:
        as_list -- return a list of the contents if true; a generator otherwise
        by_row -- iterate by row if true; by column otherwise
        """

        if by_row:
            data = self.sheet.iter_rows(self.min_col, self.max_col, self.min_row, self.max_row, values_only=True)
        else:
            data = self.sheet.iter_cols(self.min_col, self.max_col, self.min_row, self.max_row, values_only=True)
        contents = (value for row_col in data for value in row_col)

        return list(contents) if as_list else contents


    def copyTo(self, start_column=None, start_row=None, sheet=None):
        """
        Copy and paste the selection to another location.

        Arguments:
        start_column -- int
        start_row -- int
        sheet -- openpyxl Worksheet object to place the selection; default is current sheet
        """

        if sheet == None:
            sheet = self.sheet

        contents = self.getContents(as_list=False)

        for col in range(self.width+1):
            for row in range(self.height+1):
                sheet.cell(column=col+start_column, row=row+start_row).value = next(contents)


    def delete(self):
        """Delete all cell values in the selection."""

        data = self.sheet.iter_cols(self.min_col, self.max_col, self.min_row, self.max_row)
        for col in data:
            for cell in col:
                cell.value = ""


    def setFillColour(self, colour):
        """Set the fill colour of the selection"""

        data = self.sheet.iter_cols(self.min_col, self.max_col, self.min_row, self.max_row)
        for col in data:
            for cell in col:
                cell.fill = openpyxl.styles.PatternFill("solid", bgColor=colour)





def main():
    pass

if __name__ == "__main__":
    main()