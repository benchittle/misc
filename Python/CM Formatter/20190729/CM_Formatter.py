#-------------------------------------------------------------------------------
# Name:        CM_Formatter
# Purpose:     Formats an Excel file containing raw conductivity meter data.
#
# Author:      Ben Chittle
#
# Created:     16-07-2019
#-------------------------------------------------------------------------------
import GUI, Excel_Tools
import openpyxl


def findBreaks(sheet, searchCol=None):
    breaks = [4]
    for row in range(5, sheet.max_row + 1):
        if sheet.cell(row=row, column=searchCol).value > sheet.cell(row=row - 1, column=searchCol).value:
            breaks.append(row)
    breaks.append(sheet.max_row)
    return breaks


def insertDistanceColumn(sheet, Setup, lineBreaks, inputCol=None, outputCol=None):
    sheet.insert_cols(2)

    # For each line, iterates over all of the rows between it and the next line
    # and writes the current displacement from the starting point.
    for line in range(len(lineBreaks) - 1):
        for row in range(lineBreaks[line], lineBreaks[line + 1]):
            if Setup.IS_LINES_EQUAL or line not in Setup.LENGTH_EXCEPTIONS.keys():
                length = Setup.TARGET_LENGTH
            else:
                length = Setup.LENGTH_EXCEPTIONS[line]

            try:
                dist = float(sheet.cell(row=row, column=inputCol).value) * (length / sheet.cell(row=lineBreaks[line], column=inputCol).value)
            except ZeroDivisionError:
                dist = 0
            sheet.cell(row=row, column=outputCol).value = dist


def isolateLines(wb, sheet, lineBreaks, headers):
    # For each line, copy that line's data to a new sheet.
    for line in range(len(lineBreaks) - 1):
        wb.create_sheet(title="Line{}".format(line))
        newSheet = wb["Line{}".format(line)]

        # Add headers for each column.
        for col, title in enumerate(headers, start=1):
            newSheet.cell(row=1, column=col).value = title

        # Copy from the start of the current line's data to the start of the
        # next line's data.
        data = Excel_Tools.Selection(sheet, min_col=1, min_row=lineBreaks[line], max_col=sheet.max_column, max_row=lineBreaks[line + 1])
        data.copyTo(start_column=1, start_row=2, sheet=newSheet)

        # Colour code certain columns
        columns = newSheet.iter_cols(min_col=6, min_row=1, max_col=9, max_row=newSheet.max_row)
        colours = ["5b9bd5", "ed7d31", "a5a5a5", "ffc000"]
        for col, colour in zip(columns, colours):
            for cell in col:
                cell.fill = openpyxl.styles.PatternFill("solid", fgColor=colour)


def createChart(sheet, dataToGraph):
    chart = openpyxl.chart.ScatterChart()
    chartCopy = openpyxl.chart.ScatterChart()
    chart.title = chartCopy.title = sheet.title

    for col in dataToGraph:
        xValues = openpyxl.chart.Reference(sheet, min_col=2, min_row=1, max_col=2, max_row=sheet.max_row)
        yValues = openpyxl.chart.Reference(sheet, min_col=col, min_row=1, max_col=col, max_row=sheet.max_row)
        series = openpyxl.chart.Series(values=yValues, xvalues=xValues, title_from_data=True)
        chart.series.append(series)
        chartCopy.series.append(series)

    return chart, chartCopy



def main():
    Setup = GUI.Setup_Menu()
    Setup.mainloop()

    if Setup.isReady:
        wb = openpyxl.load_workbook(Setup.FILE_PATH)  ## MAKE THESE VARIABLES BETTER?? / STOP DUPLICATES
        mainSheet = wb[wb.get_sheet_names()[0]]
        mainSheet.title = "Data"

        headers = ["Cycle", "Dist (m)", "", "Line", "", "HC", "HI", "PC", "PI", "Day", "Hour", "Min", "Sec", 'S.V', 'T-K', 'CN', 'Ptch', 'Roll', 'GPS_TIME', 'Lat', 'Lon', 'GPS_A', 'lt.DMode']
        dataToGraph = [inputCol for name, inputCol in [("HC", 6), ("HI", 7), ("PC", 8), ("PI", 9)] if Setup.INCLUDED_DATA[name] == True]  # OPTIMIZE THIS

        if Setup.INCLUDED_DATA["isOmitExtra"]:
            mainSheet.delete_cols(13, 10)
            headers = headers[:13]

        lineBreaks = findBreaks(mainSheet, searchCol=3)
        insertDistanceColumn(mainSheet, Setup, lineBreaks, inputCol=1, outputCol=2)
        isolateLines(wb, mainSheet, lineBreaks, headers)

        wb.create_sheet(title="Charts")
        chartSheet = wb["Charts"]
        _chartPlaceCols = ["A", "J", "S"]
        _chartPlaceRows = [(i // 3) * 15 + 1 for i in range(len(wb.get_sheet_names()) - 2)]
        for pos, sheetName in enumerate(wb.get_sheet_names()[1:-1]):
            sheet = wb[sheetName]
            chart, chartCopy = createChart(sheet, dataToGraph)

            sheet.add_chart(chart, "{}1".format(openpyxl.utils.get_column_letter(sheet.max_column + 1)))
            chartSheet.add_chart(chartCopy, "{}{}".format(_chartPlaceCols[pos % 3], _chartPlaceRows[pos]))

        wb.save("test.xlsx")
        print("Formatted")


    print("Done")

if __name__ == '__main__':
    main()
"5b9bd5"