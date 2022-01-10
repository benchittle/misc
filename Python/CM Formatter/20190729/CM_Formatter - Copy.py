#-------------------------------------------------------------------------------
# Name:        CM_Formatter
# Purpose:
#
# Author:      BenPc
#
# Created:     16-07-2019
# Copyright:   (c) BenPc 2019
# Licence:     <your licence>
#-------------------------------------------------------------------------------
import GUI
import openpyxl as excel

class CM_Formatter():
      def __init__(self, filePath, includedData, targetLength, isLinesEqual, lengthExceptions):
          self.filePath = filePath
          self.includedData = includedData
          self.targetLength = targetLength
          self.isLinesEqual = isLinesEqual
          self.lengthExceptions = lengthExceptions

          self.wb = excel.load_workbook(filePath)
          self.sheets = self.wb.get_sheet_names()

          self.mainSheet = self.wb.get_sheet_by_name(self.sheets[0])
          self.mainTotalRows = self.mainSheet.max_row
          self.mainLineBreaks = list()
          self.columnsToCopy = range(1, 14) if self.includedData["isOmitExtra"] else range(1, 25)

          self.lengthCode = self.mainSheet.cell(row=2, column=6).value[-1]


      def deleteExtraData(self):
          self.wb.create_sheet(title="Data")
          newSheet = self.wb.get_sheet_by_name("Data")

          self.copyPasteColumns(self.mainSheet, newSheet, range(4, self.mainTotalRows + 1), self.columnsToCopy[:1])
          self.copyPasteColumns(self.mainSheet, newSheet, range(4, self.mainTotalRows + 1), self.columnsToCopy[1:-1], newColOffset=1)

          self.wb.remove_sheet(self.mainSheet)
          self.mainSheet = newSheet


      def copyPasteColumns(self, sheet, newSheet, rows, columns, newRowOffset=0, newColOffset=0):
          for col in columns:
              for row in rows:
                  newSheet.cell(row=row + newRowOffset, column=col + newColOffset).value = self.mainSheet.cell(row=row, column=col).value


      def getLineBreaks(self, col=4):
          self.mainLineBreaks.append(4)

          for row in range(col + 1, self.mainTotalRows + 1):
              if self.mainSheet.cell(row=row, column=col).value > self.mainSheet.cell(row=row - 1, column=col).value:
                 self.mainLineBreaks.append(row)

          self.mainLineBreaks.append(self.mainTotalRows)


      def seperateLines(self):
          for i in range(1, len(self.mainLineBreaks)):
              self.wb.create_sheet(title="Line{}".format(i - 1))
              newSheet = self.wb.get_sheet_by_name("Line{}".format(i - 1))

              self.copyPasteColumns(self.mainSheet, newSheet, range(self.mainLineBreaks[i - 1], self.mainLineBreaks[i]), self.columnsToCopy, 4 - self.mainLineBreaks[i - 1])


      def determineDistances(self, col=2):
          for line in range(1, len(self.mainLineBreaks)):
              for row in range(self.mainLineBreaks[line - 1], self.mainLineBreaks[line]):

                  if self.isLinesEqual or line - 1 not in self.lengthExceptions.keys():
                      length = self.targetLength
                  else:
                      length = self.lengthExceptions[line - 1]

                  self.mainSheet.cell(row=row, column=2).value = float(self.mainSheet.cell(row=row, column=1).value) * (length / self.mainSheet.cell(row=self.mainLineBreaks[line] - 1, column=1).value)


      def setTitles(self, row=3):
          colTitles = ["Cycle", "Dist (m)", "", "Line", "", "HC", "HI", "PC", "PI", "Day", "Hour", "Min", "Sec"]

          for i in self.wb.get_sheet_names():
              sheet = self.wb.get_sheet_by_name(i)

              sheet.cell(row=1, column=1).value = "File: {}".format(self.filePath)
              sheet.cell(row=2, column=1).value = "Sensor Length Code: {}".format(self.lengthCode)
              for col in range(1, len(colTitles) + 1):
                  sheet.cell(row=row, column=col).value = colTitles[col - 1]


      def drawCharts(self): #Optimize to just "Draw chart function" -> be less specific
          chartColumnData = [(data, col) for data, col in [("HC", 6), ("HI", 7), ("PC", 8), ("PI", 9)] if self.includedData[data] == True]
          self.wb.create_sheet(title="Charts")
          chartSheet = self.wb.get_sheet_by_name("Charts")

          lineSheetNames = self.wb.get_sheet_names()[1:-1]

          chartCoord = "N1" if self.includedData["isOmitExtra"] else "X1"

          placeCopyColumns = ["A", "J", "S"]
          placeCopyRows = [(i // 3) * 15 + 1 for i in range(len(lineSheetNames))]

          for lineNum, sheetName in enumerate(lineSheetNames):
              sheet = self.wb.get_sheet_by_name(sheetName)
              lastRow = sheet.max_row

              chart = excel.chart.ScatterChart()
              chart.title = "Line {}".format(lineNum)
              chartCopy = excel.chart.ScatterChart()
              chartCopy.title = "Line {}".format(lineNum)

              for data, col in chartColumnData:
                  xValues = excel.chart.Reference(sheet, min_col=2, min_row=4, max_col=2, max_row=lastRow)
                  yValues = excel.chart.Reference(sheet, min_col=col, min_row=3, max_col=col, max_row=lastRow)
                  series = excel.chart.Series(values=yValues, xvalues=xValues, title_from_data=True)

                  chart.series.append(series)
                  chartCopy.series.append(series)

              sheet.add_chart(chart, chartCoord)
              chartSheet.add_chart(chartCopy, "{}{}".format(placeCopyColumns[lineNum % 3], placeCopyRows[lineNum]))



def main():
    Setup = GUI.Setup_Menu()
    Setup.mainloop()

    if Setup.isReady:
        Formatter = CM_Formatter(Setup.FILE_PATH, Setup.INCLUDED_DATA, Setup.TARGET_LENGTH, Setup.IS_LINES_EQUAL, Setup.LENGTH_EXCEPTIONS)
        Formatter.deleteExtraData()
        Formatter.getLineBreaks() ##
        Formatter.determineDistances()
        Formatter.seperateLines()
        Formatter.setTitles()
        Formatter.drawCharts()


        Formatter.wb.save("test.xlsx")
        print("Formatted")


    print("Done")

if __name__ == '__main__':
    main()
