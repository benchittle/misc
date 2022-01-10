#-------------------------------------------------------------------------------
# Name:        CMFormatter
# Purpose:     Formats an Excel file containing raw conductivity meter data.
#
# Author:      Ben Chittle
#
# Created:     16-07-2019
# Updated      08-08-2019
#-------------------------------------------------------------------------------
import os, re
import openpyxl
import GUI, Excel_Tools


def find_breaks(sheet, searchCol=None):

    # Iterates over the 'line' column and flags the current row whenever a new
    # line starts.
    breaks = [4]
    for row in range(5, sheet.max_row + 1):
        if (sheet.cell(row=row, column=searchCol).value > sheet.cell(row=row - 1, column=searchCol).value):
            breaks.append(row)
    breaks.append(sheet.max_row)
    return breaks


def insert_distance_column(sheet, Setup, lineBreaks, inputCol=None, outputCol=None):

    sheet.insert_cols(2)

    # For each line, iterates over all of the rows between it and the next line
    # and writes the current displacement from the starting point.
    for line in range(len(lineBreaks) - 1):
        totalCycles = int(
            sheet.cell(row=lineBreaks[line+1] - 1, column=inputCol).value
            )
        for row in range(lineBreaks[line], lineBreaks[line + 1] + 1):
            if Setup.IS_LINES_EQUAL or line not in Setup.LENGTH_EXCEPTIONS.keys():
                length = Setup.TARGET_LENGTH
            else:
                length = Setup.LENGTH_EXCEPTIONS[line]

            cycle = int(sheet.cell(row=row, column=inputCol).value)
            dist = cycle * (length / totalCycles)
            sheet.cell(row=row, column=outputCol).value = dist


def isolate_lines(wb, sheet, lineBreaks, headers):

    # For each line, copy that line's data to a new sheet.
    for line in range(len(lineBreaks) - 1):
        wb.create_sheet(title="Line{}".format(line))
        newSheet = wb["Line{}".format(line)]

        # Add headers for each column.
        for col, title in enumerate(headers, start=1):
            newSheet.cell(row=1, column=col).value = title

        # Copy from the start of the current line's data to the start of the
        # next line's data.
        data = Excel_Tools.Selection(
            sheet,
            min_col=1,
            min_row=lineBreaks[line],
            max_col=sheet.max_column,
            max_row=lineBreaks[line + 1] - 1
            )
        data.copyTo(start_column=1, start_row=2, sheet=newSheet)

        # Colour code the data columns.
        columns = newSheet.iter_cols(
            min_col=6,
            min_row=1,
            max_col=9,
            max_row=newSheet.max_row
            )
        colours = ["5b9bd5", "ed7d31", "a5a5a5", "ffc000"]
        for col, colour in zip(columns, colours):
            for cell in col:
                cell.fill = openpyxl.styles.PatternFill("solid", fgColor=colour)


def createChart(sheet, dataToGraph):

    chart = openpyxl.chart.ScatterChart()
    chartCopy = openpyxl.chart.ScatterChart()
    chart.title = chartCopy.title = sheet.title

    for col in dataToGraph:
        xValues = openpyxl.chart.Reference(
            sheet,
            min_col=2,
            min_row=2,
            max_col=2,
            max_row=sheet.max_row
            )
        yValues = openpyxl.chart.Reference(
            sheet,
            min_col=col,
            min_row=1,
            max_col=col,
            max_row=sheet.max_row
            )
        series = openpyxl.chart.Series(
            values=yValues,
            xvalues=xValues,
            title_from_data=True
            )
        chart.series.append(series)
        chartCopy.series.append(series)

    return chart, chartCopy



def main():

    print("Starting GUI...")
    Setup = GUI.SetupMenu()
    print("Ready")
    Setup.mainloop()

    if Setup.isReady:
        print("Setup completed. Beginning formatting...")
        isSuccess = False

        try:
            # Workbook dnd 'data' sheet
            print("\tOpening workbook...")
            wb = openpyxl.load_workbook(Setup.FILE_PATH)  ## MAKE THESE VARIABLES BETTER?? / STOP DUPLICATES
            mainSheet = wb[wb.get_sheet_names()[0]]
            mainSheet.title = "Data"

            headers = [
                "Cycle",
                "Dist (m)",
                "",
                "Line",
                "",
                "HC",
                "HI",
                "PC",
                "PI",
                "Day",
                "Hour",
                "Min",
                "Sec",
                'S.V',
                'T-K',
                'CN',
                'Ptch',
                'Roll',
                'GPS_TIME',
                'Lat',
                'Lon',
                'GPS_A',
                'lt.DMode'
                ]
            dataToGraph = [
                inputCol
                for name, inputCol in [("HC", 6), ("HI", 7), ("PC", 8), ("PI", 9)]
                if Setup.INCLUDED_DATA[name] == True
                ]  # OPTIMIZE THIS

            # 'data' sheet and line sheets
            if Setup.INCLUDED_DATA["isOmitExtra"]:
                print("\t\tOmitting extra data...")
                mainSheet.delete_cols(13, 10)
                headers = headers[:13]
            print("\tGetting line breaks...")
            lineBreaks = find_breaks(mainSheet, searchCol=3)
            print("\tInserting distance column...")
            insert_distance_column(
                mainSheet,
                Setup,
                lineBreaks,
                inputCol=1,
                outputCol=2
                )
            print("\tCopying each line's data to a new sheet...")
            isolate_lines(wb, mainSheet, lineBreaks, headers)

            # Charts
            if len(dataToGraph) > 0:
                print("\tCreating charts...")
                wb.create_sheet(title="Charts")
                chartSheet = wb["Charts"]
                print("\t\tDetermining chart final positions...")
                _chartPlaceCols = ["A", "J", "S"]
                _chartPlaceRows = [(i // 3) * 15 + 1 for i in range(len(wb.get_sheet_names()) - 2)]
                for pos, sheetName in enumerate(wb.get_sheet_names()[1:-1]):
                    sheet = wb[sheetName]
                    print("\t\tAdding data to charts...")
                    chart, chartCopy = createChart(sheet, dataToGraph)
                    print("\t\tApplying chart to sheet '{}'...".format(sheetName))
                    sheet.add_chart(chart, "{}1".format(
                        openpyxl.utils.get_column_letter(sheet.max_column + 1)
                        ))
                    print("\t\tApplying chart to sheet 'Charts'...")
                    chartSheet.add_chart(chartCopy, "{}{}".format(
                        _chartPlaceCols[pos % 3], _chartPlaceRows[pos]
                        ))

            # Closing
            print("\tGetting output path...")
            finalPath = "{}_CMED.xlsx".format(
                os.path.splitext(Setup.FILE_PATH)[0]
                )
            print("\tSaving...")
            wb.save(finalPath)
            isSuccess = True

        finally:
            print("=" * 80)
            if isSuccess:
               print("Format successful.\nFile saved to {}".format(finalPath))
            else:
                print("Format failed.\nMake sure the file being saved to wasn't open and that nothing was changed in the original file.")
            print("=" * 80)

    else:
        print("Format did not start as the setup was not completed.")

    print("Done")
    input("\n\nPress Enter to continue...")

if __name__ == '__main__':
    main()